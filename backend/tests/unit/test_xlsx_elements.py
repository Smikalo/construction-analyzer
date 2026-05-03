"""Tests for XLSX workbook summary extraction helpers."""

from __future__ import annotations

import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table

from app.services.xlsx_elements import extract_xlsx


def _save_workbook(workbook: Workbook, tmp_path: Path, filename: str) -> Path:
    path = tmp_path / filename
    workbook.save(path)
    return path


def _patch_cached_formula_value(path: Path, cell_ref: str, cached_value: object) -> None:
    sheet_xml = "xl/worksheets/sheet1.xml"
    pattern = rf'(<c r="{cell_ref}"><f>.*?</f><v>)(.*?)(</v></c>)'

    with zipfile.ZipFile(path, "r") as source:
        entries = {name: source.read(name) for name in source.namelist()}

    xml = entries[sheet_xml].decode("utf-8")
    updated_xml, count = re.subn(
        pattern,
        lambda match: f"{match.group(1)}{cached_value}{match.group(3)}",
        xml,
        count=1,
    )
    assert count == 1
    entries[sheet_xml] = updated_xml.encode("utf-8")

    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as target:
        for name, data in entries.items():
            target.writestr(name, data)


def test_extract_xlsx_emits_workbook_and_sheet_summaries(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    visible_sheet = workbook.active
    visible_sheet.title = "Visible Sheet"
    visible_sheet["A1"] = "Alpha"
    visible_sheet["A2"] = 2
    visible_sheet["B1"] = "=SUM(A1:A2)"
    visible_sheet["B2"].comment = Comment("Needs review", "Planner")

    hidden_sheet = workbook.create_sheet("Hidden Sheet")
    hidden_sheet["A1"].comment = Comment("Hidden note", "Planner")
    hidden_sheet["C3"] = "Hidden value"
    hidden_sheet.sheet_state = "hidden"

    path = _save_workbook(workbook, tmp_path, "workbook.xlsx")

    elements = extract_xlsx(str(path), source="workbook.xlsx", document_id="xlsx-1")
    summary_elements = elements[:3]

    assert [element.element_type for element in summary_elements] == [
        "file_summary",
        "sheet_summary",
        "sheet_summary",
    ]
    assert [element.extraction_mode for element in summary_elements] == [
        "xlsx_summary",
        "xlsx_sheet_summary",
        "xlsx_sheet_summary",
    ]
    assert all(element.document_id == "xlsx-1" for element in summary_elements)
    assert all(element.source == "workbook.xlsx" for element in summary_elements)
    assert all(element.path == str(path) for element in summary_elements)
    assert all(element.confidence == 1.0 for element in summary_elements)
    assert all(element.warnings == () for element in summary_elements)

    summary = summary_elements[0]
    assert summary.content == (
        "Subject: engineering_workbook\n"
        "Sheets: Visible Sheet, Hidden Sheet\n"
        "Visible sheets: 1\n"
        "Hidden sheets: 1\n"
        "Non-empty cells: 6\n"
        "Formula cells: 1\n"
        "Comment cells: 2"
    )
    assert summary.metadata["subject"] == "engineering_workbook"
    assert summary.metadata["sheet_count"] == 2
    assert summary.metadata["xlsx_sheets"] == ["Visible Sheet", "Hidden Sheet"]
    assert summary.metadata["xlsx_visible_sheet_count"] == 1
    assert summary.metadata["xlsx_hidden_sheet_count"] == 1
    assert summary.metadata["xlsx_non_empty_cell_count"] == 6
    assert summary.metadata["xlsx_formula_cell_count"] == 1
    assert summary.metadata["xlsx_comment_count"] == 2

    visible_summary = summary_elements[1]
    assert visible_summary.content == (
        "Subject: engineering_workbook\n"
        "Sheet: Visible Sheet\n"
        "Range: A1:B2\n"
        "State: visible\n"
        "Non-empty cells: 4\n"
        "Formula cells: 1\n"
        "Comment cells: 1"
    )
    assert visible_summary.metadata["subject"] == "engineering_workbook"
    assert visible_summary.metadata["xlsx_sheet"] == "Visible Sheet"
    assert visible_summary.metadata["xlsx_range"] == "A1:B2"
    assert visible_summary.metadata["xlsx_sheet_state"] == "visible"
    assert visible_summary.metadata["xlsx_non_empty_cell_count"] == 4
    assert visible_summary.metadata["xlsx_formula_cell_count"] == 1
    assert visible_summary.metadata["xlsx_comment_count"] == 1

    hidden_summary = summary_elements[2]
    assert hidden_summary.content == (
        "Subject: engineering_workbook\n"
        "Sheet: Hidden Sheet\n"
        "Range: A1:C3\n"
        "State: hidden\n"
        "Non-empty cells: 2\n"
        "Formula cells: 0\n"
        "Comment cells: 1"
    )
    assert hidden_summary.metadata["subject"] == "engineering_workbook"
    assert hidden_summary.metadata["xlsx_sheet"] == "Hidden Sheet"
    assert hidden_summary.metadata["xlsx_range"] == "A1:C3"
    assert hidden_summary.metadata["xlsx_sheet_state"] == "hidden"
    assert hidden_summary.metadata["xlsx_non_empty_cell_count"] == 2
    assert hidden_summary.metadata["xlsx_formula_cell_count"] == 0
    assert hidden_summary.metadata["xlsx_comment_count"] == 1


def test_extract_xlsx_emits_literal_formula_comment_and_missing_cache_facts(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Calculation Sheet"
    sheet["A1"] = "Alpha"
    sheet["B1"] = "=SUM(1, 2)"
    sheet["C1"].comment = Comment("Needs review", "Planner")
    sheet["D1"] = "=AVERAGE(4, 6)"

    path = _save_workbook(workbook, tmp_path, "calculation.xlsx")
    _patch_cached_formula_value(path, "B1", 3)

    elements = extract_xlsx(str(path), source="calculation.xlsx", document_id="xlsx-2")
    granular_elements = elements[2:]

    assert [element.element_type for element in granular_elements] == [
        "cell",
        "formula",
        "comment",
        "formula",
    ]

    literal, cached_formula, comment, missing_formula = granular_elements

    assert literal.content == "Sheet: Calculation Sheet\nCell: A1\nValue: Alpha"
    assert literal.metadata == {
        "subject": "engineering_workbook",
        "xlsx_sheet": "Calculation Sheet",
        "xlsx_cell": "A1",
        "xlsx_sheet_state": "visible",
        "xlsx_value": "Alpha",
        "xlsx_value_kind": "literal",
        "extraction_certainty": "exact",
    }
    assert literal.warnings == ()

    assert cached_formula.content == (
        "Sheet: Calculation Sheet\nCell: B1\nFormula: =SUM(1, 2)\nCached value: 3"
    )
    assert cached_formula.metadata == {
        "subject": "engineering_workbook",
        "xlsx_sheet": "Calculation Sheet",
        "xlsx_cell": "B1",
        "xlsx_sheet_state": "visible",
        "xlsx_row_label": "Alpha",
        "xlsx_label": "Alpha",
        "xlsx_formula": "=SUM(1, 2)",
        "xlsx_cached_value": 3,
        "xlsx_value_kind": "cached_formula_result",
        "extraction_certainty": "exact_formula_cached_value",
    }
    assert cached_formula.warnings == ()

    assert comment.content == (
        "Sheet: Calculation Sheet\nCell: C1\nComment by Planner: Needs review"
    )
    assert comment.metadata == {
        "subject": "engineering_workbook",
        "xlsx_sheet": "Calculation Sheet",
        "xlsx_cell": "C1",
        "xlsx_sheet_state": "visible",
        "xlsx_comment_author": "Planner",
        "xlsx_comment_text": "Needs review",
        "extraction_certainty": "exact",
    }
    assert comment.warnings == ()

    assert missing_formula.content == (
        "Sheet: Calculation Sheet\nCell: D1\nFormula: =AVERAGE(4, 6)\nCached value: <missing>"
    )
    assert missing_formula.metadata == {
        "subject": "engineering_workbook",
        "xlsx_sheet": "Calculation Sheet",
        "xlsx_cell": "D1",
        "xlsx_sheet_state": "visible",
        "xlsx_row_label": "Alpha",
        "xlsx_label": "Alpha",
        "xlsx_formula": "=AVERAGE(4, 6)",
        "xlsx_value_kind": "missing_cached_value",
        "extraction_certainty": "exact_formula_cached_value_unknown",
    }
    assert missing_formula.warnings == ("missing_cached_value",)


def test_extract_xlsx_preserves_hidden_sheet_comment_provenance_and_skips_blank_cells(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    visible_sheet = workbook.active
    visible_sheet.title = "Visible"
    visible_sheet["A1"] = "Visible value"

    hidden_sheet = workbook.create_sheet("Hidden")
    hidden_sheet["B2"].comment = Comment("Hidden note", "Planner")
    hidden_sheet.sheet_state = "hidden"

    path = _save_workbook(workbook, tmp_path, "hidden.xlsx")

    elements = extract_xlsx(str(path), source="hidden.xlsx", document_id="xlsx-3")
    hidden_elements = [
        element for element in elements[3:] if element.metadata.get("xlsx_sheet") == "Hidden"
    ]

    assert [element.element_type for element in hidden_elements] == ["comment"]
    hidden_comment = hidden_elements[0]
    assert hidden_comment.content == ("Sheet: Hidden\nCell: B2\nComment by Planner: Hidden note")
    assert hidden_comment.metadata == {
        "subject": "engineering_workbook",
        "xlsx_sheet": "Hidden",
        "xlsx_cell": "B2",
        "xlsx_sheet_state": "hidden",
        "xlsx_comment_author": "Planner",
        "xlsx_comment_text": "Hidden note",
        "extraction_certainty": "exact",
    }
    assert hidden_comment.warnings == ()
    assert not any(
        element.metadata.get("xlsx_sheet") == "Hidden" and element.element_type == "cell"
        for element in elements
    )
    assert not any(
        element.metadata.get("xlsx_sheet") == "Hidden" and element.element_type == "formula"
        for element in elements
    )


def test_extract_xlsx_emits_table_and_named_range_facts(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    sheet["A1"] = "Item"
    sheet["B1"] = "Value"
    sheet["A2"] = "A"
    sheet["B2"] = 1
    sheet["A3"] = "B"
    sheet["B3"] = 2
    sheet.add_table(Table(displayName="SpecTable", ref="A1:B3"))

    sheet["D1"] = "Name"
    sheet["E1"] = "Amount"
    sheet["D2"] = "North"
    sheet["E2"] = 5
    workbook.defined_names.add(DefinedName("RangeBlock", attr_text="Sheet1!$D$1:$E$2"))

    path = _save_workbook(workbook, tmp_path, "factbook.xlsx")

    elements = extract_xlsx(str(path), source="factbook.xlsx", document_id="xlsx-4")
    facts = [element for element in elements if element.element_type in {"range", "table"}]

    assert [element.element_type for element in facts] == ["range", "table"]

    named_range, table = facts
    assert named_range.extraction_mode == "xlsx_range"
    assert named_range.content == ("| Name | Amount |\n| --- | --- |\n| North | 5 |")
    assert named_range.confidence == 1.0
    assert named_range.warnings == ()
    assert named_range.metadata == {
        "table_rows": 2,
        "table_columns": 2,
        "subject": "engineering_workbook",
        "xlsx_sheet": "Sheet1",
        "xlsx_range": "D1:E2",
        "xlsx_range_name": "RangeBlock",
        "xlsx_sheet_state": "visible",
    }

    assert table.extraction_mode == "xlsx_table"
    assert table.content == ("| Item | Value |\n| --- | --- |\n| A | 1 |\n| B | 2 |")
    assert table.confidence == 1.0
    assert table.warnings == ()
    assert table.metadata == {
        "table_rows": 3,
        "table_columns": 2,
        "subject": "engineering_workbook",
        "xlsx_sheet": "Sheet1",
        "xlsx_range": "A1:B3",
        "xlsx_table_name": "SpecTable",
        "xlsx_sheet_state": "visible",
    }


def test_extract_xlsx_adds_row_and_column_labels_and_units(tmp_path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Labels"

    sheet["A1"] = "Region"
    sheet["B1"] = "Load [kN]"
    sheet["A2"] = "North [kN]"
    sheet["B2"] = 12

    path = _save_workbook(workbook, tmp_path, "labels.xlsx")

    elements = extract_xlsx(str(path), source="labels.xlsx", document_id="xlsx-5")
    cell = next(
        element
        for element in elements
        if element.element_type == "cell" and element.metadata.get("xlsx_cell") == "B2"
    )

    assert cell.metadata == {
        "subject": "engineering_workbook",
        "xlsx_sheet": "Labels",
        "xlsx_cell": "B2",
        "xlsx_sheet_state": "visible",
        "xlsx_row_label": "North [kN]",
        "xlsx_column_label": "Load [kN]",
        "xlsx_label": "North [kN]",
        "xlsx_unit": "kN",
        "xlsx_value": 12,
        "xlsx_value_kind": "literal",
        "extraction_certainty": "exact",
    }
    assert cell.warnings == ()


def test_extract_xlsx_marks_ambiguous_units_and_unsupported_named_ranges(
    tmp_path: Path,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Conflict"

    sheet["A1"] = "Header"
    sheet["B1"] = "Load [kN]"
    sheet["A2"] = "Length [m]"
    sheet["B2"] = 5
    workbook.defined_names.add(
        DefinedName("DynamicRange", attr_text="OFFSET(Conflict!$A$1,0,0,2,1)")
    )
    workbook.defined_names.add(DefinedName("ExternalRange", attr_text="[Book2.xlsx]Conflict!$A$1"))

    path = _save_workbook(workbook, tmp_path, "conflict.xlsx")

    elements = extract_xlsx(str(path), source="conflict.xlsx", document_id="xlsx-6")
    summary = next(element for element in elements if element.element_type == "file_summary")
    cell = next(
        element
        for element in elements
        if element.element_type == "cell" and element.metadata.get("xlsx_cell") == "B2"
    )

    assert summary.warnings == ("unsupported_named_range",)
    assert summary.metadata["xlsx_unsupported_named_range_count"] == 2
    assert not any(element.element_type == "range" for element in elements)
    assert cell.metadata == {
        "subject": "engineering_workbook",
        "xlsx_sheet": "Conflict",
        "xlsx_cell": "B2",
        "xlsx_sheet_state": "visible",
        "xlsx_row_label": "Length [m]",
        "xlsx_column_label": "Load [kN]",
        "xlsx_label": "Length [m]",
        "xlsx_unit_candidates": ["m", "kN"],
        "xlsx_value": 5,
        "xlsx_value_kind": "literal",
        "extraction_certainty": "exact",
    }
    assert cell.warnings == ("ambiguous_unit",)
