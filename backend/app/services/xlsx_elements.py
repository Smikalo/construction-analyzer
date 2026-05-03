"""Helpers for normalizing XLSX workbook evidence into typed document elements."""

from __future__ import annotations

import re
from dataclasses import dataclass, replace
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]
from openpyxl.utils.cell import get_column_letter, range_boundaries  # type: ignore[import-untyped]
from openpyxl.workbook.workbook import Workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from app.services.document_elements import DocumentElement

from . import table_elements

XLSX_ELEMENT_TYPE = "file_summary"
XLSX_SHEET_ELEMENT_TYPE = "sheet_summary"
XLSX_CELL_ELEMENT_TYPE = "cell"
XLSX_FORMULA_ELEMENT_TYPE = "formula"
XLSX_COMMENT_ELEMENT_TYPE = "comment"
XLSX_SUMMARY_MODE = "xlsx_summary"
XLSX_SHEET_SUMMARY_MODE = "xlsx_sheet_summary"
XLSX_CELL_MODE = "xlsx_cell"
XLSX_FORMULA_MODE = "xlsx_formula"
XLSX_COMMENT_MODE = "xlsx_comment"
XLSX_TABLE_MODE = "xlsx_table"
XLSX_RANGE_MODE = "xlsx_range"
XLSX_RANGE_ELEMENT_TYPE = "range"
XLSX_SUBJECT = "engineering_workbook"
XLSX_LITERAL_VALUE_KIND = "literal"
XLSX_CACHED_FORMULA_VALUE_KIND = "cached_formula_result"
XLSX_MISSING_CACHED_VALUE_KIND = "missing_cached_value"
XLSX_EXACT_CERTAINTY = "exact"
XLSX_EXACT_FORMULA_CACHED_VALUE_CERTAINTY = "exact_formula_cached_value"
XLSX_EXACT_FORMULA_CACHED_VALUE_UNKNOWN_CERTAINTY = "exact_formula_cached_value_unknown"
XLSX_MISSING_CACHED_VALUE_WARNING = "missing_cached_value"
XLSX_AMBIGUOUS_UNIT_WARNING = "ambiguous_unit"
XLSX_UNSUPPORTED_NAMED_RANGE_WARNING = "unsupported_named_range"

_LABEL_UNIT_PATTERN = re.compile(r"^(?P<label>.*?)(?:\s*[\(\[](?P<unit>[^()\[\]]+)[\)\]])\s*$")


@dataclass(frozen=True, slots=True)
class XlsxSheetStats:
    """Cell-level counts for a single worksheet."""

    non_empty_cell_count: int
    formula_cell_count: int
    comment_count: int


@dataclass(frozen=True, slots=True)
class XlsxWorkbookStats:
    """Workbook-level counts derived from worksheet scans."""

    visible_sheet_count: int
    hidden_sheet_count: int
    non_empty_cell_count: int
    formula_cell_count: int
    comment_count: int


def extract_xlsx(
    path: str,
    *,
    source: str,
    document_id: str | None = None,
) -> list[DocumentElement]:
    """Extract a workbook summary, sheet summaries, and granular cell facts from XLSX."""
    formula_workbook = load_workbook(path, data_only=False, read_only=False)
    try:
        cached_workbook = load_workbook(path, data_only=True, read_only=False)
        try:
            workbook_stats, sheet_stats = _summarize_workbook(formula_workbook)
            cached_sheets = {worksheet.title: worksheet for worksheet in cached_workbook.worksheets}
            sheet_names = [worksheet.title for worksheet in formula_workbook.worksheets]
            named_range_elements, unsupported_named_ranges = _extract_named_range_elements(
                formula_workbook,
                document_id=document_id,
                source=source,
                path=path,
            )

            summary_metadata = {
                "subject": XLSX_SUBJECT,
                "sheet_count": len(sheet_names),
                "xlsx_sheets": sheet_names,
                "xlsx_visible_sheet_count": workbook_stats.visible_sheet_count,
                "xlsx_hidden_sheet_count": workbook_stats.hidden_sheet_count,
                "xlsx_non_empty_cell_count": workbook_stats.non_empty_cell_count,
                "xlsx_formula_cell_count": workbook_stats.formula_cell_count,
                "xlsx_comment_count": workbook_stats.comment_count,
            }
            if unsupported_named_ranges:
                summary_metadata["xlsx_unsupported_named_range_count"] = len(
                    unsupported_named_ranges
                )

            elements = [
                DocumentElement(
                    document_id=document_id,
                    source=source,
                    path=path,
                    page=None,
                    element_type=XLSX_ELEMENT_TYPE,
                    extraction_mode=XLSX_SUMMARY_MODE,
                    content=_render_workbook_summary(sheet_names, workbook_stats),
                    confidence=1.0,
                    warnings=(XLSX_UNSUPPORTED_NAMED_RANGE_WARNING,)
                    if unsupported_named_ranges
                    else (),
                    metadata=summary_metadata,
                )
            ]

            for worksheet, stats in zip(formula_workbook.worksheets, sheet_stats, strict=True):
                elements.append(
                    DocumentElement(
                        document_id=document_id,
                        source=source,
                        path=path,
                        page=None,
                        element_type=XLSX_SHEET_ELEMENT_TYPE,
                        extraction_mode=XLSX_SHEET_SUMMARY_MODE,
                        content=_render_sheet_summary(worksheet, stats),
                        confidence=1.0,
                        warnings=(),
                        metadata={
                            "subject": XLSX_SUBJECT,
                            "xlsx_sheet": worksheet.title,
                            "xlsx_range": worksheet.calculate_dimension(),
                            "xlsx_sheet_state": worksheet.sheet_state,
                            "xlsx_non_empty_cell_count": stats.non_empty_cell_count,
                            "xlsx_formula_cell_count": stats.formula_cell_count,
                            "xlsx_comment_count": stats.comment_count,
                        },
                    )
                )

            elements.extend(named_range_elements)
            for worksheet in formula_workbook.worksheets:
                elements.extend(
                    _extract_sheet_elements(
                        formula_workbook,
                        worksheet,
                        cached_sheets[worksheet.title],
                        document_id=document_id,
                        source=source,
                        path=path,
                    )
                )

            return elements
        finally:
            cached_workbook.close()
    finally:
        formula_workbook.close()


def _extract_sheet_elements(
    workbook: Workbook,
    worksheet: Worksheet,
    cached_worksheet: Worksheet,
    *,
    document_id: str | None,
    source: str,
    path: str,
) -> list[DocumentElement]:
    elements: list[DocumentElement] = []
    for table in sorted(worksheet.tables.values(), key=lambda table: (table.name or "").lower()):
        table_element = _build_table_element(
            worksheet,
            table,
            document_id=document_id,
            source=source,
            path=path,
        )
        if table_element is not None:
            elements.append(table_element)

    last_non_empty_label_by_column: dict[int, str] = {}
    for row in worksheet.iter_rows():
        last_non_empty_label_left: str | None = None
        for cell in row:
            row_label = last_non_empty_label_left
            column_label = last_non_empty_label_by_column.get(cell.column)
            if _is_formula_cell(cell):
                elements.append(
                    _build_formula_element(
                        worksheet,
                        cached_worksheet[cell.coordinate],
                        cell,
                        document_id=document_id,
                        source=source,
                        path=path,
                        row_label=row_label,
                        column_label=column_label,
                    )
                )
            elif cell.value is not None:
                elements.append(
                    _build_literal_element(
                        worksheet,
                        cell,
                        document_id=document_id,
                        source=source,
                        path=path,
                        row_label=row_label,
                        column_label=column_label,
                    )
                )

            if cell.comment is not None:
                elements.append(
                    _build_comment_element(
                        worksheet,
                        cell,
                        document_id=document_id,
                        source=source,
                        path=path,
                    )
                )

            label_text = _cell_label_text(cell)
            if label_text is not None:
                last_non_empty_label_left = label_text
                last_non_empty_label_by_column[cell.column] = label_text
    return elements


def _build_literal_element(
    worksheet: Worksheet,
    cell: Cell,
    *,
    document_id: str | None,
    source: str,
    path: str,
    row_label: str | None = None,
    column_label: str | None = None,
) -> DocumentElement:
    metadata = _base_cell_metadata(worksheet, cell)
    context_metadata, context_warnings = _cell_context_metadata(
        row_label=row_label,
        column_label=column_label,
    )
    metadata.update(context_metadata)
    metadata.update(
        {
            "xlsx_value": cell.value,
            "xlsx_value_kind": XLSX_LITERAL_VALUE_KIND,
            "extraction_certainty": XLSX_EXACT_CERTAINTY,
        }
    )
    return DocumentElement(
        document_id=document_id,
        source=source,
        path=path,
        page=None,
        element_type=XLSX_CELL_ELEMENT_TYPE,
        extraction_mode=XLSX_CELL_MODE,
        content=_render_literal_content(worksheet, cell),
        confidence=1.0,
        warnings=context_warnings,
        metadata=metadata,
    )


def _build_formula_element(
    worksheet: Worksheet,
    cached_cell: Cell,
    cell: Cell,
    *,
    document_id: str | None,
    source: str,
    path: str,
    row_label: str | None = None,
    column_label: str | None = None,
) -> DocumentElement:
    formula = str(cell.value)
    cached_value = cached_cell.value
    metadata = _base_cell_metadata(worksheet, cell)
    context_metadata, context_warnings = _cell_context_metadata(
        row_label=row_label,
        column_label=column_label,
    )
    metadata.update(context_metadata)
    metadata.update(
        {
            "xlsx_formula": formula,
            "xlsx_sheet_state": worksheet.sheet_state,
        }
    )

    if cached_value is None:
        warnings = (XLSX_MISSING_CACHED_VALUE_WARNING, *context_warnings)
        metadata.update(
            {
                "xlsx_value_kind": XLSX_MISSING_CACHED_VALUE_KIND,
                "extraction_certainty": XLSX_EXACT_FORMULA_CACHED_VALUE_UNKNOWN_CERTAINTY,
            }
        )
    else:
        warnings = context_warnings
        metadata.update(
            {
                "xlsx_cached_value": cached_value,
                "xlsx_value_kind": XLSX_CACHED_FORMULA_VALUE_KIND,
                "extraction_certainty": XLSX_EXACT_FORMULA_CACHED_VALUE_CERTAINTY,
            }
        )

    return DocumentElement(
        document_id=document_id,
        source=source,
        path=path,
        page=None,
        element_type=XLSX_FORMULA_ELEMENT_TYPE,
        extraction_mode=XLSX_FORMULA_MODE,
        content=_render_formula_content(worksheet, cell, cached_value),
        confidence=1.0,
        warnings=warnings,
        metadata=metadata,
    )


def _build_comment_element(
    worksheet: Worksheet,
    cell: Cell,
    *,
    document_id: str | None,
    source: str,
    path: str,
) -> DocumentElement:
    comment = cell.comment
    assert comment is not None
    metadata = _base_cell_metadata(worksheet, cell)
    metadata.update(
        {
            "xlsx_comment_author": comment.author or "",
            "xlsx_comment_text": comment.text or "",
            "extraction_certainty": XLSX_EXACT_CERTAINTY,
        }
    )
    return DocumentElement(
        document_id=document_id,
        source=source,
        path=path,
        page=None,
        element_type=XLSX_COMMENT_ELEMENT_TYPE,
        extraction_mode=XLSX_COMMENT_MODE,
        content=_render_comment_content(worksheet, cell),
        confidence=1.0,
        warnings=(),
        metadata=metadata,
    )


def _base_cell_metadata(worksheet: Worksheet, cell: Cell) -> dict[str, Any]:
    return {
        "subject": XLSX_SUBJECT,
        "xlsx_sheet": worksheet.title,
        "xlsx_cell": cell.coordinate,
        "xlsx_sheet_state": worksheet.sheet_state,
    }


def _extract_named_range_elements(
    workbook: Workbook,
    *,
    document_id: str | None,
    source: str,
    path: str,
) -> tuple[list[DocumentElement], list[str]]:
    elements: list[DocumentElement] = []
    unsupported_named_range_names: set[str] = set()

    for defined_name in sorted(
        workbook.defined_names.values(),
        key=lambda defined_name: (defined_name.name or "").lower(),
    ):
        destinations = list(defined_name.destinations)
        if not destinations:
            if defined_name.name:
                unsupported_named_range_names.add(defined_name.name)
            continue

        for sheet_name, cell_range in destinations:
            if sheet_name not in workbook.sheetnames:
                if defined_name.name:
                    unsupported_named_range_names.add(defined_name.name)
                continue

            range_element = _build_range_element(
                workbook[sheet_name],
                defined_name.name or cell_range,
                cell_range,
                document_id=document_id,
                source=source,
                path=path,
            )
            if range_element is not None:
                elements.append(range_element)

    return elements, sorted(unsupported_named_range_names, key=str.lower)


def _build_table_element(
    worksheet: Worksheet,
    table: Any,
    *,
    document_id: str | None,
    source: str,
    path: str,
) -> DocumentElement | None:
    rows = _rows_from_range(worksheet, table.ref)
    element = table_elements.table_element_from_rows(
        rows,
        document_id=document_id,
        source=source,
        path=path,
        page=None,
        confidence=1.0,
        metadata={
            "subject": XLSX_SUBJECT,
            "xlsx_sheet": worksheet.title,
            "xlsx_range": _normalize_range_reference(table.ref),
            "xlsx_table_name": table.name,
            "xlsx_sheet_state": worksheet.sheet_state,
        },
    )
    if element is None:
        return None
    return replace(element, extraction_mode=XLSX_TABLE_MODE)


def _build_range_element(
    worksheet: Worksheet,
    range_name: str,
    cell_range: str,
    *,
    document_id: str | None,
    source: str,
    path: str,
) -> DocumentElement | None:
    rows = _rows_from_range(worksheet, cell_range)
    element = table_elements.table_element_from_rows(
        rows,
        document_id=document_id,
        source=source,
        path=path,
        page=None,
        confidence=1.0,
        metadata={
            "subject": XLSX_SUBJECT,
            "xlsx_sheet": worksheet.title,
            "xlsx_range": _normalize_range_reference(cell_range),
            "xlsx_range_name": range_name,
            "xlsx_sheet_state": worksheet.sheet_state,
        },
    )
    if element is None:
        return None
    return replace(
        element,
        element_type=XLSX_RANGE_ELEMENT_TYPE,
        extraction_mode=XLSX_RANGE_MODE,
    )


def _rows_from_range(worksheet: Worksheet, cell_range: str) -> list[list[object | None]]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except ValueError:
        return []

    return [
        [cell.value for cell in row]
        for row in worksheet.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
        )
    ]


def _normalize_range_reference(cell_range: str) -> str:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except ValueError:
        return cell_range
    start = f"{get_column_letter(min_col)}{min_row}"
    end = f"{get_column_letter(max_col)}{max_row}"
    if start == end:
        return start
    return f"{start}:{end}"


def _cell_context_metadata(
    *,
    row_label: str | None,
    column_label: str | None,
) -> tuple[dict[str, Any], tuple[str, ...]]:
    metadata: dict[str, Any] = {}
    warnings: list[str] = []

    if row_label is not None:
        metadata["xlsx_row_label"] = row_label
    if column_label is not None:
        metadata["xlsx_column_label"] = column_label

    preferred_label = row_label or column_label
    if preferred_label is not None:
        metadata["xlsx_label"] = preferred_label

    row_unit = _extract_explicit_unit(row_label)
    column_unit = _extract_explicit_unit(column_label)
    if row_unit is not None and column_unit is not None:
        if row_unit == column_unit:
            metadata["xlsx_unit"] = row_unit
        else:
            metadata["xlsx_unit_candidates"] = [row_unit, column_unit]
            warnings.append(XLSX_AMBIGUOUS_UNIT_WARNING)
    elif row_unit is not None:
        metadata["xlsx_unit"] = row_unit
    elif column_unit is not None:
        metadata["xlsx_unit"] = column_unit

    return metadata, tuple(warnings)


def _cell_label_text(cell: Cell) -> str | None:
    if cell.value is None or _is_formula_cell(cell):
        return None
    text = _normalize_label_text(cell.value)
    return text or None


def _normalize_label_text(value: object) -> str:
    return " ".join(str(value).split()).strip()


def _extract_explicit_unit(label: str | None) -> str | None:
    if label is None:
        return None
    match = _LABEL_UNIT_PATTERN.match(label)
    if match is None:
        return None
    unit = match.group("unit").strip()
    if not unit or not any(ch.isalpha() for ch in unit):
        return None
    if not match.group("label").strip():
        return None
    return unit


def _render_literal_content(worksheet: Worksheet, cell: Cell) -> str:
    return "\n".join(
        [
            f"Sheet: {worksheet.title}",
            f"Cell: {cell.coordinate}",
            f"Value: {cell.value}",
        ]
    )


def _render_formula_content(
    worksheet: Worksheet,
    cell: Cell,
    cached_value: object | None,
) -> str:
    cached_text = "<missing>" if cached_value is None else str(cached_value)
    return "\n".join(
        [
            f"Sheet: {worksheet.title}",
            f"Cell: {cell.coordinate}",
            f"Formula: {cell.value}",
            f"Cached value: {cached_text}",
        ]
    )


def _render_comment_content(worksheet: Worksheet, cell: Cell) -> str:
    comment = cell.comment
    assert comment is not None
    author = comment.author or ""
    text = comment.text or ""
    return "\n".join(
        [
            f"Sheet: {worksheet.title}",
            f"Cell: {cell.coordinate}",
            f"Comment by {author}: {text}",
        ]
    )


def _summarize_workbook(workbook: Workbook) -> tuple[XlsxWorkbookStats, list[XlsxSheetStats]]:
    sheet_stats: list[XlsxSheetStats] = []
    visible_sheet_count = 0
    hidden_sheet_count = 0
    non_empty_cell_count = 0
    formula_cell_count = 0
    comment_cell_count = 0

    for worksheet in workbook.worksheets:
        stats = _summarize_sheet(worksheet)
        sheet_stats.append(stats)
        if _is_visible_sheet(worksheet):
            visible_sheet_count += 1
        else:
            hidden_sheet_count += 1
        non_empty_cell_count += stats.non_empty_cell_count
        formula_cell_count += stats.formula_cell_count
        comment_cell_count += stats.comment_count

    return (
        XlsxWorkbookStats(
            visible_sheet_count=visible_sheet_count,
            hidden_sheet_count=hidden_sheet_count,
            non_empty_cell_count=non_empty_cell_count,
            formula_cell_count=formula_cell_count,
            comment_count=comment_cell_count,
        ),
        sheet_stats,
    )


def _summarize_sheet(worksheet: Worksheet) -> XlsxSheetStats:
    non_empty_cell_count = 0
    formula_cell_count = 0
    comment_cell_count = 0

    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None or cell.comment is not None:
                non_empty_cell_count += 1
            if _is_formula_cell(cell):
                formula_cell_count += 1
            if cell.comment is not None:
                comment_cell_count += 1

    return XlsxSheetStats(
        non_empty_cell_count=non_empty_cell_count,
        formula_cell_count=formula_cell_count,
        comment_count=comment_cell_count,
    )


def _render_workbook_summary(sheet_names: list[str], stats: XlsxWorkbookStats) -> str:
    return "\n".join(
        [
            f"Subject: {XLSX_SUBJECT}",
            f"Sheets: {', '.join(sheet_names)}",
            f"Visible sheets: {stats.visible_sheet_count}",
            f"Hidden sheets: {stats.hidden_sheet_count}",
            f"Non-empty cells: {stats.non_empty_cell_count}",
            f"Formula cells: {stats.formula_cell_count}",
            f"Comment cells: {stats.comment_count}",
        ]
    )


def _render_sheet_summary(worksheet: Worksheet, stats: XlsxSheetStats) -> str:
    return "\n".join(
        [
            f"Subject: {XLSX_SUBJECT}",
            f"Sheet: {worksheet.title}",
            f"Range: {worksheet.calculate_dimension()}",
            f"State: {worksheet.sheet_state}",
            f"Non-empty cells: {stats.non_empty_cell_count}",
            f"Formula cells: {stats.formula_cell_count}",
            f"Comment cells: {stats.comment_count}",
        ]
    )


def _is_visible_sheet(worksheet: Worksheet) -> bool:
    return worksheet.sheet_state == "visible"


def _is_formula_cell(cell: Cell) -> bool:
    return cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("="))


__all__ = [
    "XLSX_ELEMENT_TYPE",
    "XLSX_SHEET_ELEMENT_TYPE",
    "XLSX_SHEET_SUMMARY_MODE",
    "XLSX_SUBJECT",
    "XLSX_SUMMARY_MODE",
    "extract_xlsx",
]
